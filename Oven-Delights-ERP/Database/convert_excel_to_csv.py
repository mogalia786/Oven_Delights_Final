from openpyxl import load_workbook
import csv

excel_path = r"C:\Development Apps\Cascades projects\Oven-Delights-ERP\Oven-Delights-ERP\Oven-Delights-ERP\Database\Copy of ITEM LIST NEW 2025 updated.xlsx"

try:
    # Load workbook
    wb = load_workbook(excel_path, data_only=True)
    print(f"Found {len(wb.sheetnames)} sheet(s): {wb.sheetnames}")
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {ws.dimensions}")
        print(f"{'='*60}")
    
        # Get all data
        data = list(ws.values)
        
        if data:
            headers = data[0]
            print(f"\nColumn headers ({len(headers)} columns):")
            for i, col in enumerate(headers, 1):
                print(f"  {i}. {col}")
            
            print(f"\nTotal rows: {len(data)}")
            print(f"\nFirst 5 data rows:")
            for i, row in enumerate(data[1:6], 1):
                print(f"  Row {i}: {row}")
            
            # Save to CSV
            csv_path = excel_path.replace('.xlsx', f'_{sheet_name}.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerows(data)
            
            print(f"\nConverted to: {csv_path}")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
